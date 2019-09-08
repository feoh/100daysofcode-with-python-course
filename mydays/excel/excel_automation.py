from openpyxl import load_workbook
from os import environ
import requests
import json


filename="nyt_popular.xlsx"
wb = load_workbook(filename)
ws = wb.create_sheet("NYT Popular Headlines")

api_key = environ['API_KEY']
popular_resp = requests.get("https://api.nytimes.com/svc/mostpopular/v2/viewed/1.json", params={'api-key':api_key})
popular_resp.raise_for_status()

popular_json = popular_resp.text
popular_dict = json.loads(popular_json)

for headline in popular_dict['results']:
    ws.append([headline['published_date'], headline['title']])

wb.save(filename=filename)




